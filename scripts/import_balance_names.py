import argparse
import json
import shutil
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parent.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))


def read_json(path):
    return json.loads(path.read_text(encoding="utf-8"))


def write_json(path, data):
    path.write_text(json.dumps(data, indent=4, ensure_ascii=False), encoding="utf-8")


def clean_name(value):
    if value is None:
        return None

    name = str(value).strip()
    if not name:
        return None

    if "\n<@" in name:
        name = name.split("\n<@", 1)[0].strip()

    return name


def import_names_from_workbook(data, workbook_path):
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    guild_id = workbook_path.stem.replace("balances_", "")
    if guild_id not in data:
        return 0

    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    try:
        user_col = headers.index("Usuario")
        id_col = headers.index("ID")
    except ValueError:
        return 0

    updated = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        user_name = clean_name(row[user_col])
        user_id = str(row[id_col]).strip() if row[id_col] is not None else ""

        if not user_name or not user_id:
            continue

        if user_id not in data[guild_id]:
            continue

        if data[guild_id][user_id].get("name") != user_name:
            data[guild_id][user_id]["name"] = user_name
            updated += 1

    return updated


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", default=".tmp_import_data/data")
    parser.add_argument("--target", default="data/balances.json")
    args = parser.parse_args()

    source = Path(args.source)
    target = Path(args.target)
    backup_dir = target.parent / "backups" / f"names_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    backup_dir.mkdir(parents=True, exist_ok=True)
    shutil.copy2(target, backup_dir / target.name)

    data = read_json(target)
    updated = 0

    for workbook_path in source.glob("balances_*.xlsx"):
        updated += import_names_from_workbook(data, workbook_path)

    write_json(target, data)

    from repositories.balance_repository import BalanceRepository

    BalanceRepository().replace_all(data)

    print(f"Names updated: {updated}")
    print(f"Backup: {backup_dir}")
    print("SQLite updated: data/bot.sqlite3")


if __name__ == "__main__":
    main()
