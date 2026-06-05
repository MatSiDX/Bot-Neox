import os

import discord
from discord.ext import commands
from discord import app_commands

from repositories.balance_repository import DATA_DIR
from services.avalonian_service import AvalonianService
from services.balance_service import BalanceService
from services.permission_service import PermissionService
from services.report_service import ReportService


class ExportCog(commands.Cog):
    def __init__(self, bot):
        self.bot = bot
        self.service = BalanceService()
        self.avalonian_service = AvalonianService()
        self.permission_service = PermissionService()
        self.report_service = ReportService()

    def has_role(self, interaction):
        return self.permission_service.can_manage_balance(interaction.guild.id, interaction.user)

    async def resolve_member_name(self, guild, user_id):
        member = guild.get_member(user_id)
        if member:
            return member.display_name

        try:
            member = await guild.fetch_member(user_id)
            return member.display_name
        except discord.NotFound:
            return f"Usuario {user_id}"
        except discord.HTTPException:
            return f"Usuario {user_id}"

    @app_commands.command(name="export")
    async def export(self, interaction: discord.Interaction):
        if not self.has_role(interaction):
            await interaction.response.send_message(
                "No tienes permisos para usar este comando.",
                ephemeral=True,
            )
            return

        await interaction.response.defer()

        ranking = self.service.get_ranking(interaction.guild)
        if not ranking:
            await interaction.followup.send("No hay datos para exportar en este servidor.")
            return

        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        os.makedirs(DATA_DIR, exist_ok=True)
        file_path = os.path.join(DATA_DIR, f"balances_{interaction.guild.id}.xlsx")
        workbook = Workbook()

        balances_sheet = workbook.active
        balances_sheet.title = "Balances"
        register_sheet = workbook.create_sheet("Registro Balance")
        avalonian_sheet = workbook.create_sheet("Registro Avas")
        report_sheet = workbook.create_sheet("Registro Informes")

        header_fill = PatternFill(start_color="234E70", end_color="234E70", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        center = Alignment(horizontal="center", vertical="center")
        left = Alignment(horizontal="left", vertical="center")
        left_wrapped = Alignment(horizontal="left", vertical="center", wrap_text=True)
        center_wrapped = Alignment(horizontal="center", vertical="center", wrap_text=True)
        add_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
        remove_fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
        items_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        silver_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        odd_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        even_fill = PatternFill(start_color="EEF3F8", end_color="EEF3F8", fill_type="solid")
        thin_side = Side(style="thin", color="D6DCE5")
        border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        balance_headers = ["Usuario", "ID", "Items", "Silver", "Total"]
        balances_sheet.append(balance_headers)

        repo_data = self.service.repo.load().get(str(interaction.guild.id), {})
        for user_id, total in ranking:
            balance = repo_data.get(str(user_id), {"items": 0, "silver": 0})
            display_name = await self.resolve_member_name(interaction.guild, user_id)
            balances_sheet.append([
                display_name,
                str(user_id),
                balance.get("items", 0),
                balance.get("silver", 0),
                total,
            ])

        register_headers = [
            "Accion",
            "Operador",
            "ID Operador",
            "Jugador",
            "ID Jugador",
            "Tipo",
            "Categoria",
            "Cantidad",
            "Balance Anterior",
            "Balance Nuevo",
            "Fecha",
            "Hora",
        ]
        register_sheet.append(register_headers)

        operations = self.service.get_operations(interaction.guild)
        for operation in operations:
            register_sheet.append([
                operation.get("action", ""),
                operation.get("operator", ""),
                operation.get("operator_id", ""),
                operation.get("player", ""),
                operation.get("player_id", ""),
                operation.get("type", ""),
                operation.get("category", ""),
                operation.get("amount", 0),
                operation.get("previous_balance", 0),
                operation.get("new_balance", 0),
                operation.get("date", ""),
                operation.get("time", ""),
            ])

        avalonian_headers = [
            "Ava N°",
            "Accion",
            "Usuario",
            "ID Usuario",
            "Cupo",
            "Justificacion",
            "Fecha",
            "Hora",
        ]
        avalonian_sheet.append(avalonian_headers)

        avalonian_interactions = self.avalonian_service.get_interactions(interaction.guild.id)
        for avalonian_interaction in avalonian_interactions:
            reason = avalonian_interaction.get("reason", "")
            if not reason and avalonian_interaction.get("action", "") == "SIGNUP":
                reason = "-"

            avalonian_sheet.append([
                avalonian_interaction.get("ava", ""),
                avalonian_interaction.get("action", ""),
                avalonian_interaction.get("user", ""),
                avalonian_interaction.get("user_id", ""),
                avalonian_interaction.get("slot", ""),
                reason,
                avalonian_interaction.get("date", ""),
                avalonian_interaction.get("time", ""),
            ])

        report_headers = [
            "Ava N°",
            "Caller",
            "ID Caller",
            "Revisado por",
            "ID",
            "Decision",
            "Motivo",
            "Fecha",
            "Hora",
        ]
        report_sheet.append(report_headers)

        report_reviews = self.report_service.get_reviews(interaction.guild.id)
        for report in report_reviews:
            report_sheet.append([
                report.get("ava", ""),
                report.get("caller", ""),
                report.get("caller_id", ""),
                report.get("reviewer", ""),
                report.get("reviewer_id", ""),
                report.get("decision", ""),
                report.get("reason", "-"),
                report.get("date", ""),
                report.get("time", ""),
            ])

        balances_widths = {
            "A": 32,
            "B": 24,
            "C": 14,
            "D": 14,
            "E": 16,
        }
        register_widths = {
            "A": 22,
            "B": 26,
            "C": 22,
            "D": 26,
            "E": 22,
            "F": 14,
            "G": 16,
            "H": 16,
            "I": 20,
            "J": 20,
            "K": 16,
            "L": 14,
        }
        avalonian_widths = {
            "A": 10,
            "B": 14,
            "C": 26,
            "D": 22,
            "E": 18,
            "F": 42,
            "G": 16,
            "H": 14,
        }
        report_widths = {
            "A": 10,
            "B": 26,
            "C": 22,
            "D": 26,
            "E": 22,
            "F": 16,
            "G": 42,
            "H": 16,
            "I": 14,
        }

        for sheet, widths in (
            (balances_sheet, balances_widths),
            (register_sheet, register_widths),
            (avalonian_sheet, avalonian_widths),
            (report_sheet, report_widths),
        ):
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
            sheet.sheet_view.showGridLines = False
            sheet.row_dimensions[1].height = 32

            for col_letter, width in widths.items():
                sheet.column_dimensions[col_letter].width = width

            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center
                cell.border = border

            for row_index in range(2, sheet.max_row + 1):
                default_height = 24 if sheet is balances_sheet else 30
                sheet.row_dimensions[row_index].height = default_height
                row_fill = odd_fill if row_index % 2 == 0 else even_fill
                for cell in sheet[row_index]:
                    cell.fill = row_fill
                    cell.border = border
                    cell.alignment = center

        for row in balances_sheet.iter_rows(min_row=2, min_col=3, max_col=5):
            for cell in row:
                cell.number_format = "#,##0"

        for row in register_sheet.iter_rows(min_row=2, min_col=8, max_col=10):
            for cell in row:
                cell.number_format = "#,##0"

        for row in register_sheet.iter_rows(min_row=2, max_col=12):
            row[0].alignment = left_wrapped
            row[1].alignment = center_wrapped
            row[3].alignment = center_wrapped
            type_cell = row[5]
            if type_cell.value == "ADD":
                type_cell.fill = add_fill
            elif type_cell.value == "REMOVE":
                type_cell.fill = remove_fill

            category_cell = row[6]
            if category_cell.value == "Items":
                category_cell.fill = items_fill
            elif category_cell.value == "Silver":
                category_cell.fill = silver_fill

        signup_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
        leave_fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
        for row in avalonian_sheet.iter_rows(min_row=2, max_col=8):
            row[1].fill = signup_fill if row[1].value == "SIGNUP" else leave_fill
            row[2].alignment = center_wrapped
            row[5].alignment = left_wrapped

        accepted_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
        rejected_fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
        for row in report_sheet.iter_rows(min_row=2, max_col=9):
            if row[5].value == "Aceptado":
                row[5].fill = accepted_fill
            elif row[5].value == "Rechazado":
                row[5].fill = rejected_fill
            row[1].alignment = center_wrapped
            row[3].alignment = center_wrapped
            row[6].alignment = left_wrapped

        workbook.save(file_path)

        await interaction.followup.send(
            "📊 Exportación completada:",
            file=discord.File(file_path),
        )


async def setup(bot):
    await bot.add_cog(ExportCog(bot))
